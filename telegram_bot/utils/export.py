"""
Export the generated schedule to an Excel file.
Usage: from utils.export import export_to_excel
"""
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.database.db import get_session
from src.bot.commands import get_schedule

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
HOURS = list(range(8, 17))

# Colour palette
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
LAB_FILL = PatternFill("solid", fgColor="C6EFCE")
CLASS_FILL = PatternFill("solid", fgColor="DDEBF7")
EMPTY_FILL = PatternFill("solid", fgColor="F2F2F2")

THIN = Side(style="thin")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_style(cell, text: str) -> None:
    cell.value = text
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def export_to_excel(
    output_path: str = "schedule.xlsx",
    batch: Optional[str] = None,
    section: Optional[str] = None,
    semester: Optional[int] = None,
) -> str:
    """
    Generate an Excel timetable and save to output_path.
    Returns the path of the saved file.
    """
    with get_session() as session:
        entries = get_schedule(session, batch=batch, section=section, semester=semester)

    # Build lookup: (batch, section, day, hour) → entry
    lookup = {}
    for e in entries:
        key = (e.batch, e.section, e.time_slot.day, e.time_slot.start_hour)
        lookup[key] = e

    # Determine which batch+section combos exist
    combos = sorted({(e.batch, e.section) for e in entries})

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    for batch_val, section_val in combos:
        sheet_name = f"{batch_val} Yr - Sec {section_val}"
        ws = wb.create_sheet(title=sheet_name)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(DAYS) + 1)
        title_cell = ws.cell(row=1, column=1)
        title_cell.value = (
            f"Haramaya University — IT Department\n"
            f"{batch_val} Year, Section {section_val}"
            + (f", Semester {semester}" if semester else "")
        )
        title_cell.font = Font(bold=True, size=13, color="1F4E79")
        title_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 40

        # Day headers (row 2)
        _header_style(ws.cell(row=2, column=1), "Time")
        for col, day in enumerate(DAYS, start=2):
            _header_style(ws.cell(row=2, column=col), day)
            ws.column_dimensions[get_column_letter(col)].width = 22

        ws.column_dimensions["A"].width = 14

        # Time slot rows
        for row_idx, hour in enumerate(HOURS, start=3):
            time_label = f"{hour:02d}:00–{hour + 1:02d}:00"
            time_cell = ws.cell(row=row_idx, column=1, value=time_label)
            time_cell.font = Font(bold=True)
            time_cell.alignment = Alignment(horizontal="center", vertical="center")
            time_cell.border = BORDER
            ws.row_dimensions[row_idx].height = 30

            for col, day in enumerate(DAYS, start=2):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = BORDER

                entry = lookup.get((batch_val, section_val, day, hour))
                if entry:
                    room_label = entry.room.name
                    instructor = entry.course.instructor or "TBA"
                    cell.value = f"{entry.course.name}\n👤 {instructor}\n🏫 {room_label}"
                    cell.fill = LAB_FILL if entry.course.is_lab else CLASS_FILL
                    cell.font = Font(size=9)
                else:
                    cell.fill = EMPTY_FILL

    wb.save(output_path)
    return output_path
